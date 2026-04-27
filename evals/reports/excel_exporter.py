# evals/reports/excel_reporter.py
import json
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.logger import get_logger

logger = get_logger()

# ─── Style constants ────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1F4E79")   # dark blue
SUMMARY_FILL  = PatternFill("solid", start_color="2E75B6")   # mid blue
ALT_ROW_FILL  = PatternFill("solid", start_color="DEEAF1")   # light blue
PASS_FILL     = PatternFill("solid", start_color="C6EFCE")   # green
FAIL_FILL     = PatternFill("solid", start_color="FFC7CE")   # red
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Arial", size=9)
BOLD_FONT     = Font(name="Arial", bold=True, size=9)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)


def _style_header_row(ws, row_num, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER


def _style_data_row(ws, row_num, col_count, alternate=False):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font      = BODY_FONT
        cell.alignment = LEFT
        cell.border    = THIN_BORDER
        if alternate:
            cell.fill = ALT_ROW_FILL


def _score_fill(score):
    """Green if score >= 0.7, red otherwise."""
    try:
        return PASS_FILL if float(score) >= 0.7 else FAIL_FILL
    except (TypeError, ValueError):
        return None


def load_all_results(reports_dir="evals/reports"):
    """Load all per-model JSON reports."""
    files = glob.glob(os.path.join(reports_dir, "report_*.json"))
    all_data = []
    for f in files:
        with open(f, "r") as fh:
            data = json.load(fh)
            all_data.extend(data)
    return all_data


def generate_excel(
    results: list = None,
    output_path: str = "evals/reports/benchmark.xlsx"
):
    """
    Generate a formatted Excel benchmark report.

    Args:
        results: list of result dicts passed directly from run_eval.py (all models combined).
                 If None, loads from JSON report files on disk.
        output_path: where to save the .xlsx file
    """
    if results is None:
        results = load_all_results()

    if not results:
        logger.warning("No results to export to Excel")
        return

    wb = Workbook()

    # ─────────────────────────────────────────────
    # SHEET 1 — Raw Data
    # ─────────────────────────────────────────────
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    ws_raw.freeze_panes = "A2"  # freeze header row

    raw_headers = [
        "Model", "Query",
        "Expected SQL", "Predicted SQL",
        "Execution Match", "Exact Match", "Partial Match", "SQL Score",
        "Agent Mode", "Agents Run", "Agents Succeeded", "Agent Score",
        "Latency (s)", "Latency Score", "Performance Score",
        "Final Score"
    ]

    ws_raw.append(raw_headers)
    _style_header_row(ws_raw, 1, len(raw_headers))

    for i, item in enumerate(results, start=2):
        sql_eval  = item.get("sql_eval", {})
        agent_eval = item.get("agent_eval", {})
        perf_eval  = item.get("performance_eval", {})
        final      = item.get("final", {})

        row = [
            item.get("model"),
            item.get("query"),
            item.get("expected_sql"),
            item.get("predicted_sql"),
            sql_eval.get("execution_match"),
            sql_eval.get("exact_match"),
            sql_eval.get("partial_match"),
            sql_eval.get("score"),
            agent_eval.get("mode"),
            agent_eval.get("agents_run", 1),
            agent_eval.get("agents_succeeded", 1),
            agent_eval.get("score"),
            perf_eval.get("latency"),
            perf_eval.get("latency_score"),
            perf_eval.get("score"),
            final.get("final_score"),
        ]

        ws_raw.append(row)
        alternate = (i % 2 == 0)
        _style_data_row(ws_raw, i, len(raw_headers), alternate)

        # colour-code final score cell
        final_cell = ws_raw.cell(row=i, column=len(raw_headers))
        fill = _score_fill(final.get("final_score"))
        if fill:
            final_cell.fill = fill
        final_cell.font = BOLD_FONT

    # column widths for raw sheet
    col_widths = [24, 45, 40, 40, 14, 12, 13, 10, 13, 11, 16, 12, 11, 13, 17, 12]
    for idx, width in enumerate(col_widths, 1):
        ws_raw.column_dimensions[get_column_letter(idx)].width = width

    ws_raw.row_dimensions[1].height = 28

    # ─────────────────────────────────────────────
    # SHEET 2 — Summary (per model averages)
    # ─────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.freeze_panes = "A2"

    sum_headers = [
        "Model", "Total Queries",
        "Avg SQL Score", "Avg Agent Score",
        "Avg Performance Score", "Avg Final Score",
        "Avg Latency (s)"
    ]

    ws_sum.append(sum_headers)
    _style_header_row(ws_sum, 1, len(sum_headers))

    # group results by model
    from collections import defaultdict
    model_groups = defaultdict(list)
    for item in results:
        model_groups[item.get("model")].append(item)

    for row_i, (model_name, items) in enumerate(model_groups.items(), start=2):
        n = len(items)

        def avg(key_path):
            vals = []
            for it in items:
                keys = key_path.split(".")
                v = it
                for k in keys:
                    v = v.get(k, {}) if isinstance(v, dict) else None
                if v is not None:
                    try:
                        vals.append(float(v))
                    except (TypeError, ValueError):
                        pass
            return round(sum(vals) / len(vals), 4) if vals else 0

        row = [
            model_name,
            n,
            avg("sql_eval.score"),
            avg("agent_eval.score"),
            avg("performance_eval.score"),
            avg("final.final_score"),
            avg("performance_eval.latency"),
        ]

        ws_sum.append(row)
        alternate = (row_i % 2 == 0)
        _style_data_row(ws_sum, row_i, len(sum_headers), alternate)

        # colour-code avg final score
        final_cell = ws_sum.cell(row=row_i, column=6)
        fill = _score_fill(row[5])
        if fill:
            final_cell.fill = fill
        final_cell.font = BOLD_FONT

    sum_widths = [28, 14, 15, 16, 20, 16, 16]
    for idx, width in enumerate(sum_widths, 1):
        ws_sum.column_dimensions[get_column_letter(idx)].width = width

    ws_sum.row_dimensions[1].height = 28

    # ─────────────────────────────────────────────
    # SHEET 3 — Multi-Agent Detail
    # Only rows where mode == "multi_agent"
    # ─────────────────────────────────────────────
    ws_ma = wb.create_sheet("Multi-Agent Detail")
    ws_ma.freeze_panes = "A2"

    ma_headers = [
        "Model", "Query",
        "Agents Run", "Agents Succeeded",
        "Planner Latency (s)", "Generator Latency (s)",
        "Agent Score", "SQL Score", "Final Score"
    ]

    ws_ma.append(ma_headers)
    _style_header_row(ws_ma, 1, len(ma_headers))

    ma_row = 2
    for item in results:
        agent_eval = item.get("agent_eval", {})
        if agent_eval.get("mode") != "multi_agent":
            continue

        latencies = agent_eval.get("agent_latencies", {})
        row = [
            item.get("model"),
            item.get("query"),
            agent_eval.get("agents_run"),
            agent_eval.get("agents_succeeded"),
            latencies.get("planner"),
            latencies.get("generator"),
            agent_eval.get("score"),
            item.get("sql_eval", {}).get("score"),
            item.get("final", {}).get("final_score"),
        ]

        ws_ma.append(row)
        alternate = (ma_row % 2 == 0)
        _style_data_row(ws_ma, ma_row, len(ma_headers), alternate)

        # colour final score
        final_cell = ws_ma.cell(row=ma_row, column=9)
        fill = _score_fill(row[8])
        if fill:
            final_cell.fill = fill
        final_cell.font = BOLD_FONT

        ma_row += 1

    ma_widths = [28, 45, 12, 17, 18, 20, 12, 11, 12]
    for idx, width in enumerate(ma_widths, 1):
        ws_ma.column_dimensions[get_column_letter(idx)].width = width

    ws_ma.row_dimensions[1].height = 28

    # ─────────────────────────────────────────────
    # Save
    # ─────────────────────────────────────────────
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logger.info(f"Excel report saved: {output_path}")
    print(f"\n✅ Excel report generated: {output_path}")